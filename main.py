from calendar import c
from codecs import replace_errors
from email import header
from fileinput import filename
from turtle import title
from weakref import finalize
from dotenv import load_dotenv
import os
import json
import token
from flask import Flask, jsonify, redirect,request,session
import urllib
from numpy import number
import requests
import time
from openpyxl import Workbook, load_workbook
myApp = Flask(__name__)
load_dotenv()
myApp.secret_key = os.getenv("SECRET_KEY")#for sessions
#import constants including app information and important urls and uri
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
REDIRECT_URI = 'http://localhost:5000/callback'
AUTH_URL = "https://accounts.spotify.com/authorize"
TOKEN_URL = "https://accounts.spotify.com/api/token"
API_BASE_URL = "https://api.spotify.com/v1/"
#will help later when I need different scope privileges dont want to have to change all the time
scopeDict = {"Images":["ugc-image-upload"],"Spotify Connect":["user-read-playback-state","user-modify-playback-state",
              "user-read-currently-playing"],"Playback":["app-remote-control","streaming"],"Playlists":["playlist-read-private",
              "playlist-read-collaborative","playlist-modify-private","playlist-modify-public"]
              ,"Follow":["user-follow-modify","user-follow-read"],"Listening History":["user-read-playback-position"
              ,"user-top-read","user-read-recently-played"],"Library":["user-library-modify",
              "user-library-read"],"Users":["user-read-email","user-read-private"],"Open Access":["user-soa-link",
              "user-soa-unlink","soa-manage-entitlements","soa-manage-partner","soa-create-partner"]}


def get_headers(sessionToken):
    headers = {
        "Authorization" : f"Bearer {sessionToken}"
    }
    return headers
def validation():
    if 'access_token' not in session:
        return redirect('/login')
    if int(time.time()) > session['expires_at']:
        return redirect('refresh_token')

#index page the page you enter in at first
@myApp.route("/")
def index():
    return "Welcome Please <a href='/login'>login</a>"#link to login page
#login page redirects to the spotify login
@myApp.route("/login")
def login():
    scope = f"{scopeDict['Users'][0]} {scopeDict['Users'][1]} {scopeDict['Spotify Connect'][2]} {scopeDict['Spotify Connect'][0]} {scopeDict['Spotify Connect'][1]} {scopeDict['Playlists'][0]}"#scope for what you want to access

    params = {#important to get the correct access token
        'client_id':CLIENT_ID,#match it with the spotify app
        'response_type' : 'code',#eventually spotify will return with code
        'scope': scope,#scope plugin
        'redirect_uri': REDIRECT_URI,#where spotify should redirect the access code
        'show_dialog':True#allow any extra dialog
    }
    auth_url = f"{AUTH_URL}?{urllib.parse.urlencode(params)}"#this is the link we will redirect to
    return redirect(auth_url)#redirect

@myApp.route("/callback")
def callback():#callback after verification on spotify
    if 'error' in request.args:#make sure we didnt get an error
        return jsonify({"error":request.args['error']})
    if 'code' in request.args:#check to see if the response has a 'code' section
        req_body = {#unpack what spotify sent us
            'code': request.args['code'],
            'grant_type':'authorization_code',#token that is extremely important
            'redirect_uri':REDIRECT_URI,
            'client_id':CLIENT_ID,
            'client_secret':CLIENT_SECRET
        }
    response = requests.post(TOKEN_URL,data=req_body)#get the info from spotify
    token_info = response.json()#json file that came back
    session['access_token'] = token_info['access_token']#save everything in a session that is important
    session['refresh_token'] = token_info['refresh_token']
    session['expires_at'] = int(time.time()) +token_info['expires_in']
    return redirect("/playlists")#eventually redirect to playlists

@myApp.route("/playlists")
def get_playlists():#now unpack what we got
    if 'access_token' not in session:#make sure we actually have the accesstoken
        return redirect('/login')#if not redirect to login
    if int(time.time()) > session['expires_at']:#make sure the token hasnt expired
        return redirect('/refresh_token')#if it has redirect
    #new header to send to spotify to get the info    
    headers = get_headers(session['access_token'])
    response = requests.get(API_BASE_URL+"me/playlists",headers=headers)#get the json with the playlists
    playlists = response.json()#get the information about the playlists
    playlistNames = [[x['name'],x['id'],x,x['tracks']['total']] for x in playlists['items']]
    listOfSongs = []
    if not os.path.exists("spotifyPlaylistData.xlsx"):
        for x in playlistNames:
            integerDivison = x[3]//100+1
            for k in range(integerDivison):
                playlistResponse = requests.get(API_BASE_URL+f"playlists/{x[1]}/tracks?offset={100*k}&limit=100",headers=headers)
                playlistData = playlistResponse.json()
                for y in playlistData['items']:
                    listOfSongs.append(y['track']['name'])
            x.append(listOfSongs)
            listOfSongs = []
    
    #print(playlistData['items'])##added this feature to check to see if contributions are working!
    if os.path.exists("spotifyPlaylistData.xlsx"):#this will be used to make sure we dont override our data in the excel fike
        workbook = load_workbook(filename="spotifyPlaylistData.xlsx")
    else:
        workbook = Workbook()#make the excel document with the sheet names as the ids of the playlist for easy tracking!
        for z in playlistNames:
            workbook.create_sheet(title=f"{z[1]}")
            for i in range(2):
                workbook[f'{z[1]}']['A1'] = 'Name'
                workbook[f'{z[1]}']['B1'] = 'Score'
        
                #workbook[f'{playlistNames[x][1]}'][f'A{z+1}'] = finalList[x][z]
        # -> sheet name
        for y in range(len(playlistNames)):
            for z in range(len(playlistNames[y][4])):
                workbook[f'{playlistNames[y][1]}'][f'A{z+1}'] = playlistNames[y][4][z]
                workbook[f'{playlistNames[y][1]}'][f'B{z+2}'] = 100
        
        workbook.save(filename="./spotifyPlaylistData.xlsx")

    return redirect("/player")#display the json file

@myApp.route('/player')
def my_player():
    if 'access_token' not in session:#make sure we actually have the accesstoken
        return redirect('/login')#if not redirect to login
    if int(time.time()) > session['expires_at']:#make sure the token hasnt expired
        return redirect('/refresh_token')#if it has redirect
    try:
        playbackStateUrl = f"{API_BASE_URL}me/player"
        playbackStateResponse = requests.get(url=playbackStateUrl,headers=get_headers(session['access_token']))
        playbackStateJson = playbackStateResponse.json()
        currentPlaylist = playbackStateJson['context']['external_urls']['spotify'].replace("https://open.spotify.com/playlist/","")
        currentSong = playbackStateJson['item']['name']
    except:
        return "Turn on Player!"
    if playbackStateJson['is_playing']:
        if os.path.exists("spotifyPlaylistData.xlsx"):
            workbook = load_workbook(filename="spotifyPlaylistData.xlsx")
            counter = 1
            for col in  workbook[currentPlaylist].iter_cols():
                for cell in col:
                    if cell.value == currentSong:
                        saveCell = cell.value
                        break;
                    else:
                        counter += 1
                if saveCell == currentSong:
                    break
        else:
            return redirect("/playlists")
        wentBack = request.args.get("prev")
        Data = [workbook[currentPlaylist][f"A{counter}"].value,int(workbook[currentPlaylist][f"B{counter}"].value),counter]
        if wentBack == '1':
            workbook[currentPlaylist][f"B{counter}"] = int(workbook[currentPlaylist][f"B{counter}"].value) + 1
            Data[1] += 1
            workbook.save(filename="./spotifyPlaylistData.xlsx")
            return redirect("/player")
        headers = get_headers(session['access_token'])
        url = f"{API_BASE_URL}me/player/currently-playing"
        response = requests.get(url=url,headers=headers)
        currently_playing = response.json()
        deviceUrl = f"{API_BASE_URL}me/player/devices"
        deviceResponse = requests.get(url=deviceUrl,headers=get_headers(session['access_token']))
        deviceJson = deviceResponse.json()
        deviceId = deviceJson['devices'][0]['id']
        playOrPause = "<a href='/pause'>Pause</a>"
        returnString = f"""
                    Currently Playing:<br>Artist: {currently_playing['item']['artists'][0]['name']} - Song: {currently_playing['item']['name']}<br>
                    Stats: {Data[1]}<br>
                    <a href='/previous'>Prev</a> {playOrPause} <a href='/skip?name={currentSong}&playlist={currentPlaylist}'>Skip</a>"""
    else:
        playOrPause = "<a href='/play'>Play</a>"
        returnString = f"Music Currently not playing! {playOrPause}"
                    
    return returnString
    #?cell={Data[2]}
    #<img src='{currently_playing['item']['album']['images'][0]['url']}'>
    #currently_playing['item']['album']['images'][0]['url']
    #{currently_playing['item']['artists'][0]['name']}
    #{currently_playing['item']['name']}


@myApp.route('/play')
def start_play():
    url = f"{API_BASE_URL}me/player/play"
    playResponse = requests.put(url=url, headers=get_headers(session['access_token']))
    return redirect("/player")


@myApp.route('/pause')
def pause():
    url = f"{API_BASE_URL}me/player/pause"
    playResponse = requests.put(url=url, headers=get_headers(session['access_token']))
    return redirect("/player")


@myApp.route("/refresh_token")
def get_refresh_token():
    if 'refresh_token' not in session:
        return redirect('/login')

    if int(time.time()) > session['expires_at']:
        req_body = {
            'grant_type' : 'refresh_token',
            'refresh_token' : session['refresh_token'],
            'client_id':CLIENT_ID,
            'client_secret': CLIENT_SECRET

        }
        response = requests.post(TOKEN_URL, data=req_body)
        new_token_info = response.json()
        session['access_token'] = new_token_info['access_token']
        session['expires_at'] = int(time.time()) + new_token_info['expires_in']
        return redirect("/playlists")

@myApp.route('/previous')#this page is used to go back to the previous song - Spotify Premium required
def previous_Song():
    url = f"{API_BASE_URL}me/player/previous"
    headers = get_headers(session['access_token'])
    response = requests.post(url=url,headers=headers)
    return redirect("/player?prev=1")


@myApp.route('/skip')#this page is used to skip to the next song - Spotify Premium required
def skip_Song():
    currentSong = request.args.get("name")
    currentPlaylist = request.args.get("playlist")
    workbook = load_workbook(filename="spotifyPlaylistData.xlsx")
    counter = 1
    for col in  workbook[currentPlaylist].iter_cols():
        for cell in col:
            if cell.value == currentSong:
                saveCell = cell.value
                break;
            else:
                counter += 1
        if saveCell == currentSong:
            break
    workbook[currentPlaylist][f"B{counter}"] = workbook[currentPlaylist][f"B{counter}"].value -1
    workbook.save(filename="./spotifyPlaylistData.xlsx")
    url = f"{API_BASE_URL}me/player/next"
    headers = get_headers(session['access_token'])
    response = requests.post(url=url,headers=headers)
    return redirect("/player")
#things to add data analytics based on previous skips
#make sure we do not make the excel file over and over 
#make the home screen a lot better
#add a way to login in with other uses and make the overall app better



if __name__ == "__main__":
    myApp.run(host="0.0.0.0",debug=True)